from __future__ import annotations

import io
import math
from datetime import datetime, timezone
from typing import Any, Dict

import pandas as pd
from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl.chart import BarChart, LineChart, RadarChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage
from backend.app.utils.session import get_session
from backend.app.utils.session import update_session

router = APIRouter()


def _json_safe(value):
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    return value


def _build_diagnostics_excel_bytes(result: Dict[str, Any], tab_key: str) -> io.BytesIO:
    def _to_float(value: Any) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    def _safe_sheet_name(name: str, used: set[str]) -> str:
        base = (name or "Sheet").replace(":", "_").replace("/", "_").replace("\\", "_")
        base = base[:31] or "Sheet"
        if base not in used:
            used.add(base)
            return base
        idx = 2
        while True:
            suffix = f"_{idx}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
            if candidate not in used:
                used.add(candidate)
                return candidate
            idx += 1

    def _write_df(writer: pd.ExcelWriter, used: set[str], name: str, rows: Any) -> str:
        if isinstance(rows, pd.DataFrame):
            df = rows.copy()
        else:
            if not rows:
                df = pd.DataFrame([{"info": "No data available"}])
            else:
                df = pd.DataFrame(rows)
        df = _json_safe(df.replace({math.inf: None, -math.inf: None}).to_dict(orient="records"))
        sheet_name = _safe_sheet_name(name, used)
        pd.DataFrame(df).to_excel(writer, index=False, sheet_name=sheet_name)
        return sheet_name

    def _add_bar_chart(
        wb,
        used: set[str],
        source_sheet: str,
        chart_sheet_title: str,
        chart_title: str,
        category_col: int,
        data_min_col: int,
        data_max_col: int,
    ):
        if source_sheet not in wb.sheetnames:
            return
        src = wb[source_sheet]
        if src.max_row <= 1:
            return
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_title
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Category"
        data = Reference(src, min_col=data_min_col, max_col=data_max_col, min_row=1, max_row=src.max_row)
        cats = Reference(src, min_col=category_col, min_row=2, max_row=src.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_chart = wb.create_sheet(_safe_sheet_name(chart_sheet_title, used))
        ws_chart.add_chart(chart, "A1")

    def _add_line_chart(
        wb,
        used: set[str],
        source_sheet: str,
        chart_sheet_title: str,
        chart_title: str,
        category_col: int,
        data_min_col: int,
        data_max_col: int,
    ):
        if source_sheet not in wb.sheetnames:
            return
        src = wb[source_sheet]
        if src.max_row <= 1:
            return
        chart = LineChart()
        chart.style = 10
        chart.title = chart_title
        chart.y_axis.title = "Value"
        chart.x_axis.title = "X"
        data = Reference(src, min_col=data_min_col, max_col=data_max_col, min_row=1, max_row=src.max_row)
        cats = Reference(src, min_col=category_col, min_row=2, max_row=src.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_chart = wb.create_sheet(_safe_sheet_name(chart_sheet_title, used))
        ws_chart.add_chart(chart, "A1")

    def _add_radar_chart(
        wb,
        used: set[str],
        source_sheet: str,
        chart_sheet_title: str,
        chart_title: str,
        category_col: int,
        data_min_col: int,
        data_max_col: int,
    ):
        if source_sheet not in wb.sheetnames:
            return
        src = wb[source_sheet]
        if src.max_row <= 1:
            return
        chart = RadarChart()
        chart.style = 10
        chart.type = "filled"
        chart.title = chart_title
        data = Reference(src, min_col=data_min_col, max_col=data_max_col, min_row=1, max_row=src.max_row)
        cats = Reference(src, min_col=category_col, min_row=2, max_row=src.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_chart = wb.create_sheet(_safe_sheet_name(chart_sheet_title, used))
        ws_chart.add_chart(chart, "A1")

    def _add_chart_image(
        wb,
        used: set[str],
        sheet_title: str,
        plotter,
    ):
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except Exception:
            return

        try:
            fig, ax = plt.subplots(figsize=(9.5, 5.2), dpi=120)
            plotter(ax, plt)
            fig.tight_layout()
            buf = io.BytesIO()
            buf.name = f"{sheet_title}.png"
            fig.savefig(buf, format="png", dpi=120)
            plt.close(fig)
            buf.seek(0)
            ws_chart = wb.create_sheet(_safe_sheet_name(sheet_title, used))
            ws_chart.add_image(OpenpyxlImage(buf), "A1")
        except Exception:
            try:
                plt.close("all")
            except Exception:
                pass

    output = io.BytesIO()
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(output) as writer:
        if tab_key == "data":
            data_drift = result.get("data_drift", {})
            datasets = result.get("datasets", {}) or {}
            target = data_drift.get("target_drift", {}) or {}
            csi = data_drift.get("feature_csi", {}) or {}
            missing = data_drift.get("missing_rate_drift", {}) or {}
            cardinality = data_drift.get("cardinality_drift", {}) or {}
            desc = data_drift.get("descriptive_stats", {}) or {}

            # Overview + target drift tables and graph data
            _write_df(writer, used_sheet_names, "Data_Overview", [{
                "training_rows": datasets.get("training_rows"),
                "dev_oos_rows": datasets.get("dev_oos_rows"),
                "new_rows": datasets.get("new_rows"),
                "target_training_rate": target.get("training_rate"),
                "target_new_rate": target.get("new_rate"),
                "target_delta_pp": target.get("delta_pp"),
            }])

            training_n = int(_to_float(datasets.get("training_rows")))
            new_n = int(_to_float(datasets.get("new_rows")))
            training_rate = _to_float(target.get("training_rate"))
            new_rate = _to_float(target.get("new_rate"))
            target_overall_sheet = _write_df(writer, used_sheet_names, "Target_Overall", [
                {
                    "sample": "training",
                    "observations": training_n,
                    "events": round(training_n * training_rate),
                    "event_rate_pct": training_rate * 100.0,
                },
                {
                    "sample": "new",
                    "observations": new_n,
                    "events": round(new_n * new_rate),
                    "event_rate_pct": new_rate * 100.0,
                    "delta_pp_vs_training": _to_float(target.get("delta_pp")),
                },
            ])

            target_breakdown = target.get("breakdown", {}) or {}
            breakdown_rows = []
            for feature, rows in target_breakdown.items():
                for row in rows or []:
                    breakdown_rows.append({
                        "feature": feature,
                        "segment": row.get("segment"),
                        "train_obs": row.get("train_obs"),
                        "new_obs": row.get("new_obs"),
                        "train_events": row.get("train_events"),
                        "new_events": row.get("new_events"),
                        "train_rate": row.get("train_rate"),
                        "new_rate": row.get("new_rate"),
                        "delta_pp": row.get("delta_pp"),
                    })
            _write_df(writer, used_sheet_names, "Target_Breakdown", breakdown_rows)

            # CSI tables + chart data sources
            csi_rows = []
            csi_distribution_rows = []
            csi_contrib_rows = []
            for feature, details in csi.items():
                details = details or {}
                csi_rows.append({
                    "feature": feature,
                    "csi": details.get("value"),
                    "severity": details.get("severity"),
                })
                d = details.get("details", {}) or {}
                categories = d.get("categories") or []
                bins = d.get("bins") or []
                train_pct = d.get("train_pct") or []
                new_pct = d.get("new_pct") or []
                contrib = d.get("contrib") or []
                labels = categories if categories else [f"B{i+1}" for i in range(max(0, len(bins) - 1))]
                for idx, label in enumerate(labels):
                    csi_distribution_rows.append({
                        "feature": feature,
                        "label": label,
                        "train_pct": _to_float(train_pct[idx] if idx < len(train_pct) else 0) * 100.0,
                        "new_pct": _to_float(new_pct[idx] if idx < len(new_pct) else 0) * 100.0,
                    })
                    csi_contrib_rows.append({
                        "feature": feature,
                        "label": label,
                        "contribution": _to_float(contrib[idx] if idx < len(contrib) else 0),
                    })
            feature_csi_sheet = _write_df(writer, used_sheet_names, "Feature_CSI", sorted(csi_rows, key=lambda r: _to_float(r.get("csi")), reverse=True))
            _write_df(writer, used_sheet_names, "CSI_Distribution_Data", csi_distribution_rows)
            _write_df(writer, used_sheet_names, "CSI_Contribution_Data", csi_contrib_rows)

            # Cardinality + missing
            missing_rows = [{"feature": f, **(vals or {})} for f, vals in missing.items()]
            _write_df(writer, used_sheet_names, "Missing_Drift", missing_rows)
            cardinality_rows = []
            for feature, vals in cardinality.items():
                vals = vals or {}
                cardinality_rows.append({
                    "feature": feature,
                    "train_categories_count": len(vals.get("train_categories") or []),
                    "new_categories_count": len(vals.get("new_categories") or []),
                    "new_only_count": len(vals.get("new_only") or []),
                    "lost_count": len(vals.get("lost") or []),
                    "new_only": ", ".join((vals.get("new_only") or [])[:50]),
                    "lost": ", ".join((vals.get("lost") or [])[:50]),
                })
            _write_df(writer, used_sheet_names, "Cardinality_Drift", cardinality_rows)

            # Descriptive raw and processed
            raw_rows = []
            proc_rows = []
            for feature, vals in (desc.get("raw", {}) or {}).items():
                vals = vals or {}
                row = {"feature": feature}
                row.update(vals.get("training", {}) or {})
                row.update({f"new_{k}": v for k, v in (vals.get("new", {}) or {}).items()})
                raw_rows.append(row)
            for feature, vals in (desc.get("processed", {}) or {}).items():
                vals = vals or {}
                row = {"feature": feature}
                row.update(vals.get("training", {}) or {})
                row.update({f"new_{k}": v for k, v in (vals.get("new", {}) or {}).items()})
                proc_rows.append(row)
            _write_df(writer, used_sheet_names, "Descriptive_Raw", raw_rows)
            _write_df(writer, used_sheet_names, "Descriptive_Processed", proc_rows)

            wb = writer.book
            _add_bar_chart(
                wb, used_sheet_names, target_overall_sheet,
                "Chart_Target_Drift", "Target Event Rate (%)", category_col=1, data_min_col=4, data_max_col=4
            )
            _add_bar_chart(
                wb, used_sheet_names, feature_csi_sheet,
                "Chart_Feature_CSI", "Feature CSI", category_col=1, data_min_col=2, data_max_col=2
            )
            _add_chart_image(
                wb,
                used_sheet_names,
                "ChartImg_Target_Drift",
                lambda ax, _plt: (
                    ax.bar(["Training", "New data"], [training_rate * 100.0, new_rate * 100.0], color=["#9CA3AF", "#FB4E0B"]),
                    ax.set_title("Target Event Rate (%)"),
                    ax.set_ylabel("Event rate (%)"),
                    ax.grid(axis="y", alpha=0.25),
                ),
            )
            top_csi = sorted(csi_rows, key=lambda r: _to_float(r.get("csi")), reverse=True)[:15]
            if top_csi:
                _add_chart_image(
                    wb,
                    used_sheet_names,
                    "ChartImg_Feature_CSI",
                    lambda ax, _plt: (
                        ax.barh([r["feature"] for r in top_csi], [_to_float(r.get("csi")) for r in top_csi], color="#FB4E0B"),
                        ax.set_title("Feature CSI"),
                        ax.set_xlabel("CSI"),
                        ax.grid(axis="x", alpha=0.25),
                    ),
                )
        elif tab_key == "concept":
            concept = result.get("concept_drift", {})
            iv = concept.get("iv", {}) or {}
            uni_gini = concept.get("univariate_gini", {}) or concept.get("univariate_auc", {}) or {}
            bivariate = concept.get("bivariate_monotonicity", {}) or {}
            iv_rows = [{"feature": f, **(vals or {})} for f, vals in iv.items()]
            gini_rows = [{"feature": f, **(vals or {})} for f, vals in uni_gini.items()]
            iv_sheet = _write_df(writer, used_sheet_names, "IV", iv_rows)
            gini_sheet = _write_df(writer, used_sheet_names, "Univariate_Gini", gini_rows)

            monotonicity_rows = []
            for feature, vals in bivariate.items():
                vals = vals or {}
                labels = vals.get("bin_labels") or []
                train_rate = vals.get("train_rate") or []
                new_rate = vals.get("new_rate") or []
                for idx, label in enumerate(labels):
                    monotonicity_rows.append({
                        "feature": feature,
                        "bin": label,
                        "train_rate": _to_float(train_rate[idx] if idx < len(train_rate) else 0),
                        "new_rate": _to_float(new_rate[idx] if idx < len(new_rate) else 0),
                    })
            _write_df(writer, used_sheet_names, "Monotonicity_Data", monotonicity_rows)
            wb = writer.book
            _add_bar_chart(
                wb, used_sheet_names, iv_sheet,
                "Chart_IV", "Information Value (Train vs New)", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_bar_chart(
                wb, used_sheet_names, gini_sheet,
                "Chart_Univariate_Gini", "Univariate Gini (Dev Validation vs New)", category_col=1, data_min_col=2, data_max_col=3
            )
            top_iv = sorted(iv_rows, key=lambda r: abs(_to_float(r.get("delta"))), reverse=True)[:15]
            if top_iv:
                _add_chart_image(
                    wb,
                    used_sheet_names,
                    "ChartImg_IV",
                    lambda ax, _plt: (
                        ax.barh([r["feature"] for r in top_iv], [_to_float(r.get("iv_train")) for r in top_iv], color="#9CA3AF", label="Train"),
                        ax.barh([r["feature"] for r in top_iv], [_to_float(r.get("iv_new")) for r in top_iv], color="#FB4E0B", alpha=0.7, label="New"),
                        ax.set_title("Information Value"),
                        ax.legend(),
                        ax.grid(axis="x", alpha=0.25),
                    ),
                )
            def _gini_val(row: Dict[str, Any], key: str, auc_key: str) -> float:
                if row.get(key) is not None:
                    return _to_float(row.get(key))
                if row.get(auc_key) is not None:
                    return 2.0 * _to_float(row.get(auc_key)) - 1.0
                return 0.0

            top_gini = sorted(
                gini_rows,
                key=lambda r: abs(_gini_val(r, "new_gini", "new_auc") - _gini_val(r, "dev_gini", "train_auc")),
                reverse=True,
            )[:15]
            if top_gini:
                _add_chart_image(
                    wb,
                    used_sheet_names,
                    "ChartImg_Univariate_Gini",
                    lambda ax, _plt: (
                        ax.barh(
                            [r["feature"] for r in top_gini],
                            [_gini_val(r, "dev_gini", "dev_auc") or _gini_val(r, "dev_gini", "train_auc") for r in top_gini],
                            color="#9CA3AF",
                            label="Dev validation",
                        ),
                        ax.barh(
                            [r["feature"] for r in top_gini],
                            [_gini_val(r, "new_gini", "new_auc") for r in top_gini],
                            color="#FB4E0B",
                            alpha=0.7,
                            label="New",
                        ),
                        ax.set_title("Univariate Gini"),
                        ax.legend(),
                        ax.grid(axis="x", alpha=0.25),
                    ),
                )
        elif tab_key == "performance":
            perf = result.get("performance_drift", {})
            interp = result.get("interpretability", {}) or {}
            summary = {
                "auc_dev": perf.get("auc_dev"),
                "auc_new": perf.get("auc_new"),
                "auc_drop_pp": perf.get("auc_drop_pp"),
                "ks_dev": perf.get("ks_dev"),
                "ks_new": perf.get("ks_new"),
                "ks_drop_pp": perf.get("ks_drop_pp"),
                "gini_dev": perf.get("gini_dev"),
                "gini_new": perf.get("gini_new"),
                "auc_pr_new": perf.get("auc_pr_new"),
                "log_loss_new": perf.get("log_loss_new"),
                "brier_new": perf.get("brier_new"),
                "score_psi": (perf.get("score_psi") or {}).get("psi"),
            }
            perf_summary_sheet = _write_df(writer, used_sheet_names, "Performance_Summary", [summary])

            # Charts and table sources used on UI
            roc_dev_rows = perf.get("roc_curve_dev") or []
            roc_new_rows = perf.get("roc_curve_new") or []
            roc_combined_rows = []
            max_roc_len = max(len(roc_dev_rows), len(roc_new_rows))
            for i in range(max_roc_len):
                dev_row = roc_dev_rows[i] if i < len(roc_dev_rows) else {}
                new_row = roc_new_rows[i] if i < len(roc_new_rows) else {}
                roc_combined_rows.append({
                    "fpr": _to_float(dev_row.get("fpr", new_row.get("fpr", 0.0))),
                    "tpr_dev": _to_float(dev_row.get("tpr", 0.0)),
                    "tpr_new": _to_float(new_row.get("tpr", 0.0)),
                })
            roc_combined_sheet = _write_df(writer, used_sheet_names, "ROC_Combined", roc_combined_rows)
            ks_sheet = _write_df(writer, used_sheet_names, "KS_Curve_New", perf.get("ks_curve_new") or [])
            calibration_sheet = _write_df(writer, used_sheet_names, "Calibration_New", perf.get("calibration_new") or [])
            _write_df(writer, used_sheet_names, "Lift_Dev", perf.get("dev_lift_table") or [])
            _write_df(writer, used_sheet_names, "Lift_New", perf.get("new_lift_table") or [])

            # Decile event-rate chart source
            decile_dev = perf.get("decile_rates_dev") or []
            decile_new = perf.get("decile_rates_new") or []
            decile_rows = []
            for idx in range(max(len(decile_dev), len(decile_new))):
                decile_rows.append({
                    "decile": f"D{idx + 1}",
                    "dev_rate": _to_float(decile_dev[idx] if idx < len(decile_dev) else 0),
                    "new_rate": _to_float(decile_new[idx] if idx < len(decile_new) else 0),
                })
            decile_sheet = _write_df(writer, used_sheet_names, "Decile_Rates", decile_rows)

            # Classification panels
            _write_df(writer, used_sheet_names, "Thresholds", [{
                "current_threshold": perf.get("classification_threshold"),
                "ks_optimal": (perf.get("thresholds") or {}).get("ks_optimal"),
                "f1_optimal": (perf.get("thresholds") or {}).get("f1_optimal"),
            }])
            _write_df(writer, used_sheet_names, "Classification_Dev", [perf.get("classification_dev") or {}])
            _write_df(writer, used_sheet_names, "Classification_New", [perf.get("classification_new") or {}])
            _write_df(writer, used_sheet_names, "ROB_Dev", [perf.get("rob_dev") or {}])
            _write_df(writer, used_sheet_names, "ROB_New", [perf.get("rob_new") or {}])

            # SHAP and PDP
            shap_dev = interp.get("shap_importance_dev") or {}
            shap_new = interp.get("shap_importance_new") or {}
            shap_features = sorted(set(shap_dev.keys()) | set(shap_new.keys()))
            shap_rows = [{
                "feature": f,
                "dev_importance": _to_float(shap_dev.get(f)),
                "new_importance": _to_float(shap_new.get(f)),
            } for f in shap_features]
            shap_sheet = _write_df(writer, used_sheet_names, "SHAP_Importance", shap_rows)
            _write_df(writer, used_sheet_names, "SHAP_Flags", [interp.get("shap_flags") or {}])

            pdp_dev = interp.get("pdp_dev") or {}
            pdp_new = interp.get("pdp_new") or {}
            pdp_rows = []
            for feature in sorted(set(pdp_dev.keys()) | set(pdp_new.keys())):
                dev = pdp_dev.get(feature) or {}
                cur = pdp_new.get(feature) or {}
                dev_x = dev.get("x") or []
                dev_y = dev.get("y") or []
                cur_x = cur.get("x") or []
                cur_y = cur.get("y") or []
                for idx in range(max(len(dev_x), len(cur_x))):
                    pdp_rows.append({
                        "feature": feature,
                        "x_dev": _to_float(dev_x[idx] if idx < len(dev_x) else None),
                        "y_dev": _to_float(dev_y[idx] if idx < len(dev_y) else None),
                        "x_new": _to_float(cur_x[idx] if idx < len(cur_x) else None),
                        "y_new": _to_float(cur_y[idx] if idx < len(cur_y) else None),
                    })
            _write_df(writer, used_sheet_names, "PDP_Data", pdp_rows)

            radar_rows = [
                {"metric": "AUC", "dev": _to_float(perf.get("auc_dev")), "new": _to_float(perf.get("auc_new"))},
                {"metric": "KS", "dev": _to_float(perf.get("ks_dev")), "new": _to_float(perf.get("ks_new"))},
                {"metric": "Gini", "dev": _to_float(perf.get("gini_dev")), "new": _to_float(perf.get("gini_new"))},
                {"metric": "AUC-PR", "dev": _to_float(perf.get("auc_pr_dev")), "new": _to_float(perf.get("auc_pr_new"))},
            ]
            radar_sheet = _write_df(writer, used_sheet_names, "Radar_Data", radar_rows)

            wb = writer.book
            _add_bar_chart(
                wb, used_sheet_names, perf_summary_sheet,
                "Chart_Perf_Summary", "Performance Summary", category_col=1, data_min_col=1, data_max_col=12
            )
            _add_line_chart(
                wb, used_sheet_names, roc_combined_sheet,
                "Chart_ROC", "ROC Curve (Dev vs New)", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_line_chart(
                wb, used_sheet_names, ks_sheet,
                "Chart_KS", "KS Curve", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_line_chart(
                wb, used_sheet_names, calibration_sheet,
                "Chart_Calibration", "Calibration Curve", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_bar_chart(
                wb, used_sheet_names, decile_sheet,
                "Chart_Decile_Rates", "Decile Event Rates", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_bar_chart(
                wb, used_sheet_names, shap_sheet,
                "Chart_SHAP", "SHAP Importance (Dev vs New)", category_col=1, data_min_col=2, data_max_col=3
            )
            _add_radar_chart(
                wb, used_sheet_names, radar_sheet,
                "Chart_Radar", "Performance Radar", category_col=1, data_min_col=2, data_max_col=3
            )
            if roc_combined_rows:
                _add_chart_image(
                    wb,
                    used_sheet_names,
                    "ChartImg_ROC",
                    lambda ax, _plt: (
                        ax.plot([_to_float(r.get("fpr")) for r in roc_combined_rows], [_to_float(r.get("tpr_dev")) for r in roc_combined_rows], color="#9CA3AF", label="Dev"),
                        ax.plot([_to_float(r.get("fpr")) for r in roc_combined_rows], [_to_float(r.get("tpr_new")) for r in roc_combined_rows], color="#FB4E0B", label="New"),
                        ax.plot([0, 1], [0, 1], linestyle="--", color="#94A3B8", alpha=0.6),
                        ax.set_title("ROC Curve"),
                        ax.set_xlabel("FPR"),
                        ax.set_ylabel("TPR"),
                        ax.legend(),
                        ax.grid(alpha=0.25),
                    ),
                )
            if decile_rows:
                _add_chart_image(
                    wb,
                    used_sheet_names,
                    "ChartImg_Decile_Rates",
                    lambda ax, _plt: (
                        ax.bar([r["decile"] for r in decile_rows], [r["dev_rate"] for r in decile_rows], color="#9CA3AF", label="Dev"),
                        ax.bar([r["decile"] for r in decile_rows], [r["new_rate"] for r in decile_rows], color="#FB4E0B", alpha=0.7, label="New"),
                        ax.set_title("Decile Event Rates"),
                        ax.set_ylabel("Rate"),
                        ax.legend(),
                        ax.grid(axis="y", alpha=0.25),
                    ),
                )
        else:
            desc = (result.get("data_drift", {}).get("descriptive_stats", {}) or {})
            raw = desc.get("raw", {})
            proc = desc.get("processed", {})
            raw_rows = []
            proc_rows = []
            for feature, vals in raw.items():
                row = {"feature": feature}
                row.update(vals.get("training", {}))
                row.update({f"new_{k}": v for k, v in vals.get("new", {}).items()})
                raw_rows.append(row)
            for feature, vals in proc.items():
                row = {"feature": feature}
                row.update(vals.get("training", {}))
                row.update({f"new_{k}": v for k, v in vals.get("new", {}).items()})
                proc_rows.append(row)
            _write_df(writer, used_sheet_names, "Raw", raw_rows)
            _write_df(writer, used_sheet_names, "Processed", proc_rows)
    output.seek(0)
    return output


@router.get("/report")
async def get_drift_report(session_id: str = Query(...)):
    session = get_session(session_id)
    if not session:
        return {"error": "Session not found"}
    result = session.get("drift_result")
    if not result:
        return {"error": "Drift diagnostics not run yet"}
    return _json_safe(result)


@router.get("/variable/{var_name}")
async def get_variable_detail(var_name: str, session_id: str = Query(...)):
    session = get_session(session_id)
    if not session:
        return {"error": "Session not found"}
    result = session.get("drift_result")
    if not result:
        return {"error": "Drift diagnostics not run yet"}

    dist = result.get("variable_distributions", {}).get(var_name, {})
    csi = result.get("csi_results", {}).get(var_name)
    iv = result.get("iv_results", {}).get(var_name)
    woe = result.get("woe_results", {}).get(var_name)

    return _json_safe({
        "variable": var_name,
        "csi": csi,
        "iv": iv,
        "woe": woe,
        "distribution": dist,
    })


@router.post("/decision")
async def save_diagnostic_decision(body: Dict[str, Any]):
    session_id = str(body.get("session_id") or "").strip()
    gate = str(body.get("gate") or "final").strip()
    selection = str(body.get("selection") or "").strip()
    rationale = str(body.get("rationale") or "").strip()
    if not session_id:
        return JSONResponse({"error": "session_id is required"}, status_code=400)
    session = get_session(session_id)
    if not session:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    if not selection:
        return JSONResponse({"error": "selection is required"}, status_code=400)
    if not rationale:
        return JSONResponse({"error": "rationale is required"}, status_code=400)

    decisions = list(session.get("drift_decisions") or [])
    record = {
        "gate": gate,
        "selection": selection,
        "rationale": rationale,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    decisions.append(record)
    payload: Dict[str, Any] = {"drift_decisions": decisions}
    if gate == "final":
        payload["selected_recalibration_action"] = selection
    update_session(session_id, payload)
    return {"ok": True, "decision": record}


@router.get("/download/{tab}")
async def download_diagnostics_report(tab: str, session_id: str = Query(...)):
    session = get_session(session_id)
    if not session:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    result = session.get("drift_result")
    if not result:
        return JSONResponse({"error": "Drift diagnostics not run yet"}, status_code=400)

    tab_key = tab.strip().lower()
    if tab_key not in {"data", "concept", "performance", "descriptive"}:
        return JSONResponse({"error": "tab must be one of: data, concept, performance, descriptive"}, status_code=400)

    output = _build_diagnostics_excel_bytes(result, tab_key)
    filename = f"diagnostics_{tab_key}_{session_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/download/{tab}")
async def download_diagnostics_report_inline(tab: str, body: Dict[str, Any]):
    tab_key = tab.strip().lower()
    if tab_key not in {"data", "concept", "performance", "descriptive"}:
        return JSONResponse({"error": "tab must be one of: data, concept, performance, descriptive"}, status_code=400)

    report = body.get("report")
    if not isinstance(report, dict):
        return JSONResponse({"error": "report object is required"}, status_code=400)

    output = _build_diagnostics_excel_bytes(report, tab_key)
    filename = f"diagnostics_{tab_key}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
